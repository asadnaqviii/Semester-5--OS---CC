import xml.etree.ElementTree as ET  #Python parser 
import openpyxl  #Use this for manipulating excel files

tree = ET.parse('compiler.xml')  #Parse XML file
root = tree.getroot() #fetch XML file

#Create new excel file
workbook = openpyxl.Workbook()
worksheet = workbook.active

#Adding the column headers
worksheet.cell(row=1, column=1).value = 'Book ID'
worksheet.cell(row=1, column=2).value = 'Author Name'
worksheet.cell(row=1, column=3).value = 'Title'
worksheet.cell(row=1, column=4).value = 'Genre'
worksheet.cell(row=1, column=5).value = 'Price'
worksheet.cell(row=1, column=6).value = 'Publish Date'
worksheet.cell(row=1, column=7).value = 'Description'

row = 2 #adding data in 2nd row
for book in root.findall('book'):
  book_id = book.get('id')
  author_name = book.find('author').text
  title = book.find('title').text
  genre = book.find('genre').text
  price = book.find('price').text
  publish_date = book.find('publish_date').text
  description = book.find('description').text

  worksheet.cell(row=row, column=1).value = book_id
  worksheet.cell(row=row, column=2).value = author_name
  worksheet.cell(row=row, column=3).value = title
  worksheet.cell(row=row, column=4).value = genre
  worksheet.cell(row=row, column=5).value = price
  worksheet.cell(row=row, column=6).value = publish_date
  worksheet.cell(row=row, column=7).value = description
  
  row += 1

file_name = input("Enter the name for the Excel file: ")  #input to ask for filename (it will be stored in the same folder as this file)

try:   
  workbook.save(file_name + '.xlsx')   #Save the Excel file
  print("Excel file successfully created!")
except Exception as e:
  print("There was an error saving the Excel file:", e)